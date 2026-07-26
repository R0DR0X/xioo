import os
import glob
import re
import traceback
import unicodedata
import difflib
import openpyxl
import pandas as pd

# ═══════════════════════════════════════════════════════════════════════════
# PARTE 1: CLASIFICADOR (de Clasificador_Veritrade.py)
# ═══════════════════════════════════════════════════════════════════════════

# ═══ DICCIONARIO DE CATEGORÍAS Y KEYWORDS (Español, Inglés y Abreviaciones) ═══
CATEGORIES = {
    "FILETE": ["FILETE", "FILETES", "FILLET", "FILLETS", "FIL", "FILE", "DARUMA", "MANTO", "MANTOS"],
    "ALAS": ["ALETA", "ALETAS", "ALA", "ALAS", "WING", "WINGS", "ALET"],
    "TENTACULO": ["TENTACULO", "TENTACULOS", "REJO", "REJOS", "BAILARINA", "BAILARINAS", "TENTACLE", "TENTACLES", "TENT", "REJ"],
    "NUCA": ["NUCA", "NUCAS", "NECK", "NECKS", "NUC"],
    "REPRODUCTOR": ["REPRODUCTOR", "REPRODUCTORES", "SEXUAL", "SEXUALES", "REPRODUCTIVE", "REP", "SEX"]
}

# ═══ EXCLUSIONES COMPLETAS (Si se encuentra la palabra se borra/invalida de inmediato) ═══
COMPLETE_EXCLUSIONS = [
    # Cortes compuestos / sub-productos o formas procesadas
    "TUBO", "ANILLA", "RECORT", "CONO", "RABA", "SLICE", "TABLETA", "MASA", "BOTON", "COLLARIN",
    "RING", "CUBO", "DADO",
    # Productos preparados / especiales
    "ENSALADA", "SAZONADO", "MARINADO", "PICADILLO", "LAMELES", "HAMBURGUESA", "EMPANIZADO", "REBOZADO",
    # Cortes / formas especiales
    "RODAJA", "STRIPS", "CHIP", "TIRAS", "COLA", "TROZO",
    # Residuos / pulpa
    "RESIDUO", "SUBPRODUCTO", "PULPA",
    # Métodos / estados excluidos
    "IQF", "SECO"
]

NON_POTA_SPECIES = [
    "MAHI", "MERLUZA", "ANCHOVETA", "ATUN", "BONITO", "CABALLA",
    "PEJERREY", "PERICO", "DORADO", "PULPO", "JIBIA", "SEPIA",
    "LANGOSTINO", "CAMARON", "ERIZO", "CONCHAS", "CHORO", "ALMEJA",
    "VIEIRA", "JUREL", "CABINZA", "CORVINA", "LENGUADO"
]

def normalize_text(text):
    """
    Normaliza el texto quitando acentos, caracteres especiales (excepto paréntesis
    para conservar tags como (R)) y unificando espacios en mayúsculas.
    """
    if not isinstance(text, str):
        return ""
    text = text.upper()
    # Separar paréntesis con espacios para evitar fusiones de palabras como REJOS(REPRODUCTORES)
    text = text.replace("(", " ( ").replace(")", " ) ")
    # Eliminar acentos/diacríticos
    text = "".join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    # Reemplazar caracteres especiales por espacios (conservando paréntesis y letras/números)
    text = re.sub(r'[^A-Z0-9\s\(\)]', ' ', text)
    # Unificar espacios múltiples
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def extract_categories_fuzzy(normalized_desc):
    """
    Extrae las categorías de producto identificadas en el texto usando fuzzy matching
    sobre cada palabra de la descripción.
    """
    words = normalized_desc.split()
    matched_cats = set()

    # Búsqueda de tags explícitos de reproductor como (R) o ( R )
    if re.search(r'\(\s*R\s*\)', normalized_desc):
        matched_cats.add("REPRODUCTOR")

    for word in words:
        # Evitar falsos positivos de palabras genéricas de exportación con la categoría REPRODUCTOR (ej: PRODUCTO, PROD)
        if word in ["PRODUCTO", "PRODUCTOS", "PROD"]:
            continue

        for cat, keywords in CATEGORIES.items():
            for kw in keywords:
                # Si la keyword es muy corta (ej: abreviaciones de 3 letras o menos), match exacto
                if len(kw) <= 3:
                    if word == kw:
                        matched_cats.add(cat)
                        break
                else:
                    # Fuzzy matching usando difflib
                    sim = difflib.SequenceMatcher(None, word, kw).ratio()
                    # Umbral de 0.80 para palabras de longitud >= 4
                    if len(word) >= 4 and sim >= 0.80:
                        matched_cats.add(cat)
                        break
    return matched_cats

def classify_product(desc_upper):
    """
    Clasifica un producto hidrobiológico de pota/calamar gigante basándose
    en su descripción comercial, aplicando el nuevo pipeline optimizado.
    """
    if not desc_upper:
        return ""

    # 1. Normalización de texto
    normalized = normalize_text(desc_upper)

    # 2. Exclusiones absolutas/completas (Fórmula Normal)
    for kw in COMPLETE_EXCLUSIONS:
        if kw in normalized:
            return ""

    for sp in NON_POTA_SPECIES:
        if sp in normalized:
            return ""

    # Exclusión condicional de PUNTA
    if "PUNTA" in normalized:
        # Caso A: "SIN PUNTA" → negación, no excluir
        has_punta_negated = "SIN PUNTA" in normalized
        # Caso B: viene acompañado de tentáculos/rejos/bailarinas → parte del tentáculo, no excluir
        has_tentacle = ("TENTACULO" in normalized or "REJO" in normalized or "BAILARINA" in normalized)
        # Caso C: viene acompañado de reproductores → ej: "PUNTA DE REPRODUCTORES", no excluir
        has_reproductor_kw = ("REPRODUCTOR" in normalized or "REPRODUCTORES" in normalized or
                              "SEXUAL" in normalized or "SEXUALES" in normalized)
        if not has_punta_negated and not has_tentacle and not has_reproductor_kw:
            return ""

    # 3. Exclusiones parciales (Entero/Entera)
    has_entera = ("ENTERA" in normalized or "ENTERO" in normalized)

    # 4. Extracción de categorías usando fuzzy matching
    matched_cats = extract_categories_fuzzy(normalized)

    # Si contiene "entera" pero NO hay ninguna palabra de producto rescatable, se va
    if has_entera and not matched_cats:
        return ""

    # 5. Detección de productos compuestos (2+ categorías diferentes)
    if len(matched_cats) >= 2:
        # Excepción especial: Tentáculo + Reproductor
        if matched_cats == {"TENTACULO", "REPRODUCTOR"}:
            # Comprobar si hay negación ("sin reproductor", "sin sexual", "sin ventosa")
            has_negation = ("SIN VENTOSA" in normalized or
                            "SIN SEXUAL" in normalized or
                            "SIN REPRODUCTOR" in normalized)
            if has_negation:
                matched_cats = {"TENTACULO"}
            else:
                matched_cats = {"REPRODUCTOR"}
        else:
            # Producto compuesto -> vacío
            return ""

    if not matched_cats:
        return ""

    category = list(matched_cats)[0]

    # 6. Clasificación térmica (Cocido vs Congelado)
    is_cooked = ("COCID" in normalized or
                 "PRECOCID" in normalized or
                 "BOILED" in normalized or
                 bool(re.search(r'\bPC\b', normalized)))

    if category == "REPRODUCTOR":
        return "REPRODUCTOR"
    elif category == "TENTACULO":
        return "TENTACULO"
    elif category == "NUCA":
        if is_cooked:
            return ""  # Nuca cocida no es válida según reglas
        return "NUCA"
    elif category == "ALAS":
        if is_cooked:
            return "ALAS COCIDAS"
        return "ALAS CONGELADAS"
    elif category == "FILETE":
        if is_cooked or "DARUMA" in normalized:
            return "FILETE COCIDO"
        return "FILETE CONGELADO"

    return ""

def process_veritrade(file_path, output_path):
    print(f"\n=====================================")
    print(f"--- PROCESANDO VERITRADE: {os.path.basename(file_path)}")

    # ── Lectura rápida con pandas (evita el escaneo lento de openpyxl) ──
    df_raw = pd.read_excel(file_path, sheet_name=0, header=5, dtype=str)
    df_raw = df_raw.dropna(how='all')

    # Columnas por índice (0-based): X=col 23, Y=col 24
    col_x = df_raw.columns[23]  # Descripción comercial completa
    col_y = df_raw.columns[24]  # Descripcion1 (más corta y específica)

    # ── 1. Agrupar por Descripcion1 ÚNICA para reducir operaciones fuzzy ──
    # Clave de clasificación = Descripcion1 (col Y), más limpia y específica
    desc1_groups = {}  # desc1_upper -> {"desc_x": primera desc X, "count": n}
    for _, row_data in df_raw.iterrows():
        desc1_raw = str(row_data[col_y]).strip() if pd.notna(row_data[col_y]) else ""
        desc_x_raw = str(row_data[col_x]).strip() if pd.notna(row_data[col_x]) else ""
        if not desc1_raw or desc1_raw.upper() == "NAN":
            continue
        key = desc1_raw.upper()
        if key not in desc1_groups:
            desc1_groups[key] = {
                "desc1_original": desc1_raw,   # texto original (mayús/minús)
                "desc_x": desc_x_raw,           # descripción comercial larga (para contexto)
                "count": 0
            }
        desc1_groups[key]["count"] += 1

    print(f"   Filas válidas: {sum(g['count'] for g in desc1_groups.values())} | "
          f"Descripcion1 únicas: {len(desc1_groups)}")

    # ── 2. Clasificar cada Descripcion1 única UNA SOLA VEZ ──
    classification_map = {}  # desc1_upper -> producto
    for desc1_upper in desc1_groups:
        classification_map[desc1_upper] = classify_product(desc1_upper)

    # ── 3. Escribir resultados al Excel con openpyxl ──
    wb = openpyxl.load_workbook(file_path)
    ws = wb[wb.sheetnames[0]]
    new_col = 'AH'
    ws[f'{new_col}6'] = 'PRODUCTO'

    total = 0
    clasificados = 0
    stats = {}

    for row in range(7, ws.max_row + 1):
        desc1_val = ws[f'Y{row}'].value
        if desc1_val is None:
            continue
        desc1_str = str(desc1_val).strip()
        if not desc1_str or desc1_str.upper() == "NAN":
            continue

        total += 1
        key = desc1_str.upper()
        producto = classification_map.get(key, "")
        ws[f'{new_col}{row}'] = producto
        if producto:
            clasificados += 1
            stats[producto] = stats.get(producto, 0) + 1

    wb.save(output_path)
    wb.close()

    no_clasificados = total - clasificados
    base_report = os.path.splitext(output_path)[0]

    # ── 4a. MD de CLASIFICADOS ──────────────────────────────────────────────
    path_clas = base_report + "_CLASIFICADOS.md"
    with open(path_clas, 'w', encoding='utf-8') as f:
        f.write("# Reporte: Productos CLASIFICADOS\n\n")
        f.write(f"Clasificador basado en **Descripcion1** (col Y). "
                f"Total clasificados: **{clasificados:,}** de {total:,} filas.\n\n")

        # Resumen estadístico
        f.write("## Resumen por Categoría\n\n")
        f.write("| Categoría | Filas | % del Total |\n")
        f.write("| :--- | ---: | ---: |\n")
        for prod, cnt in sorted(stats.items(), key=lambda x: -x[1]):
            f.write(f"| **{prod}** | {cnt:,} | {(cnt/total)*100:.2f}% |\n")
        f.write(f"| **TOTAL CLASIFICADOS** | {clasificados:,} | {(clasificados/total)*100:.2f}% |\n\n")

        # Tabla de correspondencia – solo clasificados
        f.write("## Detalle: Descripcion1 Clasificadas (por frecuencia)\n\n")
        f.write("| Repeticiones | Descripcion1 | Clasificación | Desc. Comercial (ref.) |\n")
        f.write("| ---: | :--- | :--- | :--- |\n")
        sorted_clas = [
            (k, v) for k, v in sorted(
                desc1_groups.items(), key=lambda x: -x[1]['count']
            ) if classification_map.get(k, "")
        ]
        for key, info in sorted_clas:
            prod = classification_map[key]
            desc1_clean = info['desc1_original'].replace("|", "\\|")
            desc_x_clean = info['desc_x'].replace("|", "\\|")[:120]
            f.write(f"| {info['count']} | {desc1_clean} | `{prod}` | {desc_x_clean} |\n")

    # ── 4b. MD de NO CLASIFICADOS ────────────────────────────────────────────
    path_no_clas = base_report + "_NO_CLASIFICADOS.md"
    with open(path_no_clas, 'w', encoding='utf-8') as f:
        f.write("# Reporte: Productos NO CLASIFICADOS\n\n")
        f.write(f"Entradas de **Descripcion1** que el clasificador dejó en blanco. "
                f"Total: **{no_clasificados:,}** filas ({(no_clasificados/total)*100:.2f}% del total).\n\n")
        f.write("> Revisar esta lista para detectar **falsos negativos** "
                "(productos de pota válidos que no fueron capturados).\n\n")

        f.write("## Detalle: Descripcion1 NO Clasificadas (por frecuencia)\n\n")
        f.write("| Repeticiones | Descripcion1 | Desc. Comercial (ref.) |\n")
        f.write("| ---: | :--- | :--- |\n")
        sorted_no_clas = [
            (k, v) for k, v in sorted(
                desc1_groups.items(), key=lambda x: -x[1]['count']
            ) if not classification_map.get(k, "")
        ]
        for key, info in sorted_no_clas:
            desc1_clean = info['desc1_original'].replace("|", "\\|")
            desc_x_clean = info['desc_x'].replace("|", "\\|")[:120]
            f.write(f"| {info['count']} | {desc1_clean} | {desc_x_clean} |\n")

    # ── Resumen en consola ───────────────────────────────────────────────────
    print(f"--- LISTO! {clasificados:,}/{total:,} filas clasificadas "
          f"({(clasificados/total)*100:.1f}%)")
    print(f"   Nueva columna '{new_col}' (PRODUCTO) agregada al Excel")
    for prod, cnt in sorted(stats.items(), key=lambda x: -x[1]):
        print(f"   * {prod}: {cnt:,}")
    print(f"   >> Clasificados    -> {os.path.basename(path_clas)}")
    print(f"   >> No clasificados -> {os.path.basename(path_no_clas)}")
    print(f"   ** Excel guardado  -> {os.path.basename(output_path)}")


# ═══════════════════════════════════════════════════════════════════════════
# PARTE 2: FILTRADOR (de veritrade_filtrador.py)
# ═══════════════════════════════════════════════════════════════════════════

def find_density_interval_price(prices):
    """
    Finds the interval of size 1.0 that contains the maximum number of transactions.
    Returns the mean price of the transactions within that interval.
    If prices is empty, returns None.
    """
    if not prices:
        return None

    sorted_prices = sorted(prices)
    best_count = -1
    best_start = None
    best_prices_in_interval = []

    for p in sorted_prices:
        # Interval is [p, p + 1.0]
        in_interval = [x for x in sorted_prices if p <= x <= p + 1.0]
        count = len(in_interval)

        # Tie breaker: if same count, take the one with higher starting price (conservative)
        if count > best_count:
            best_count = count
            best_start = p
            best_prices_in_interval = in_interval
        elif count == best_count:
            if best_start is None or p > best_start:
                best_start = p
                best_prices_in_interval = in_interval

    if not best_prices_in_interval:
        return None
    return sum(best_prices_in_interval) / len(best_prices_in_interval)

def find_density_interval_price_with_pf(all_prices, pf_prices):
    """
    Finds the interval that contains all Peru Frost prices and has the highest density of all prices.
    The window size is W = max(1.0, max(pf_prices) - min(pf_prices)).
    If pf_prices is empty, falls back to standard interval search on all_prices.
    """
    if not pf_prices:
        return find_density_interval_price(all_prices)

    pf_min = min(pf_prices)
    pf_max = max(pf_prices)
    w = max(1.0, pf_max - pf_min)

    # We want to find a window [start, start + w] that covers all pf_prices.
    # Therefore, start must satisfy: pf_max - w <= start <= pf_min.
    # This range of starting points is exactly [pf_max - w, pf_min].
    lower_bound_start = pf_max - w
    upper_bound_start = pf_min

    candidates = [lower_bound_start, upper_bound_start]
    for p in all_prices:
        if lower_bound_start <= p <= upper_bound_start:
            candidates.append(p)

    candidates = sorted(list(set(candidates)))

    best_count = -1
    best_start = None
    best_prices_in_interval = []

    for start in candidates:
        # Window is [start, start + w]
        in_interval = [x for x in all_prices if start <= x <= start + w]
        count = len(in_interval)

        if count > best_count:
            best_count = count
            best_start = start
            best_prices_in_interval = in_interval
        elif count == best_count:
            if best_start is None or start > best_start:
                best_start = start
                best_prices_in_interval = in_interval

    if not best_prices_in_interval:
        return None
    return sum(best_prices_in_interval) / len(best_prices_in_interval)

def save_filtered_excel(source_excel_path, df_filtered, original_cols, output_path):
    print(f"--- Guardando en: {os.path.basename(output_path)}...")
    df_to_save = df_filtered[original_cols].copy()

    wb = openpyxl.load_workbook(source_excel_path)
    if 'Veritrade' in wb.sheetnames:
        ws = wb['Veritrade']

        header_map = {}
        for c in range(1, ws.max_column + 1):
            h_val = ws.cell(row=6, column=c).value
            if h_val:
                header_map[str(h_val).strip()] = c

        max_r = ws.max_row
        max_c = ws.max_column
        if max_r >= 7:
            for r in range(7, max_r + 1):
                for c in range(1, max_c + 1):
                    ws.cell(r, c).value = None

        rows_data = df_to_save.to_dict('records')
        for r_idx, row_dict in enumerate(rows_data, 7):
            for col_name, val in row_dict.items():
                if col_name in header_map:
                    c_idx = header_map[col_name]
                    if isinstance(val, pd.Timestamp):
                        val = val.to_pydatetime()
                    elif pd.isna(val):
                        val = None
                    ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(output_path)
    wb.close()
    print(f"--- Guardado completo!")

def process_filtering(file_path, output_path):
    print(f"\n--- INICIANDO PROCESO DE FILTRADO Y VARIACIONES: {os.path.basename(file_path)}")

    # 1. Cargar archivo (fila 6 son los encabezados, index 5)
    df_full = pd.read_excel(file_path, sheet_name='Veritrade', skiprows=5)

    # Guardar lista de columnas originales para el guardado final
    original_cols = df_full.columns.tolist()

    # Columnas necesarias
    col_exportador = 'Exportador'
    col_kg_neto = 'Kg Neto'
    col_fob_tot = 'U$ FOB Tot'
    col_producto = 'PRODUCTO'
    col_fecha = 'Fecha'

    required_cols = [col_exportador, col_kg_neto, col_fob_tot, col_producto, col_fecha]
    for col in required_cols:
        if col not in df_full.columns:
            print(f"--- ERROR: No se encontro la columna '{col}'.")
            return

    # 2. Preparar datos para el calculo
    df_full[col_fecha] = pd.to_datetime(df_full[col_fecha], errors='coerce')
    df_full = df_full.dropna(subset=[col_fecha]) # Eliminar filas sin fecha
    df_full[col_producto] = df_full[col_producto].astype(str).str.strip() # LIMPIEZA DE PRODUCTO
    df_full['MES_ANIO'] = df_full[col_fecha].dt.to_period('M')
    df_full['Precio_Fila'] = df_full[col_fob_tot] / df_full[col_kg_neto]

    target_exporter = "PERU FROST SOCIEDAD ANONIMA CERRADA"

    # Definir la ruta de la carpeta "filtrados veritrade"
    base_dir = os.path.dirname(os.path.abspath(__file__))
    folder_variations = os.path.join(base_dir, "filtrados veritrade")
    os.makedirs(folder_variations, exist_ok=True)

    # ----------------------------------------------------
    # 1. CÁLCULO PARA EL FLUJO BASE (Salida principal en FILTERED_OUTPUT)
    # ----------------------------------------------------
    df_pf = df_full[df_full[col_exportador] == target_exporter].copy()

    pf_stats = pd.DataFrame()
    if not df_pf.empty:
        pf_raw = df_pf.groupby(['MES_ANIO', col_producto]).agg({
            col_fob_tot: 'sum', col_kg_neto: 'sum'
        }).reset_index()
        pf_raw['Prom_Crudo'] = pf_raw[col_fob_tot] / pf_raw[col_kg_neto]

        df_pf = df_pf.merge(pf_raw[['MES_ANIO', col_producto, 'Prom_Crudo']], on=['MES_ANIO', col_producto], how='left')
        mask_pf_clean = df_pf['Precio_Fila'] <= (df_pf['Prom_Crudo'] + 1.2)
        df_pf_clean = df_pf[mask_pf_clean].copy()

        pf_stats = df_pf_clean.groupby(['MES_ANIO', col_producto]).agg({
            col_fob_tot: 'sum', col_kg_neto: 'sum'
        }).reset_index()
        pf_stats['Promedio_PF'] = pf_stats[col_fob_tot] / pf_stats[col_kg_neto]

    # Calcular limites mensuales para el flujo Base
    limites_base = {}
    for (mes, prod), group in df_full.groupby(['MES_ANIO', col_producto]):
        has_pf = False
        limite = None

        if not pf_stats.empty:
            pf_match = pf_stats[(pf_stats['MES_ANIO'] == mes) & (pf_stats[col_producto] == prod)]
            if not pf_match.empty:
                has_pf = True
                limite = pf_match.iloc[0]['Promedio_PF'] + 0.5

        if not has_pf:
            precios_mercado = group['Precio_Fila'].tolist()
            precio_ref = find_density_interval_price(precios_mercado)
            if precio_ref is not None:
                limite = precio_ref + 0.5

        if limite is not None:
            limites_base[(mes, prod)] = limite

    df_base = df_full.copy()
    df_base['Limite_Filtro'] = df_base.apply(lambda r: limites_base.get((r['MES_ANIO'], r[col_producto])), axis=1)

    mask_delete_base = (df_base['Limite_Filtro'].notna()) & (df_base['Precio_Fila'] > df_base['Limite_Filtro'])
    df_filtered_base = df_base[~mask_delete_base].copy()

    # Guardar archivo base
    save_filtered_excel(file_path, df_filtered_base, original_cols, output_path)

    # ----------------------------------------------------
    # 2. GENERAR VARIACIÓN 1 (Todos los meses con intervalos, nunca elimina Peru Frost)
    # ----------------------------------------------------
    limites_v1 = {}
    for (mes, prod), group in df_full.groupby(['MES_ANIO', col_producto]):
        precios_mercado = group['Precio_Fila'].tolist()
        precio_ref = find_density_interval_price(precios_mercado)
        if precio_ref is not None:
            limites_v1[(mes, prod)] = precio_ref + 0.5

    df_v1 = df_full.copy()
    df_v1['Limite_Filtro'] = df_v1.apply(lambda r: limites_v1.get((r['MES_ANIO'], r[col_producto])), axis=1)

    mask_delete_v1 = (df_v1['Limite_Filtro'].notna()) & \
                     (df_v1['Precio_Fila'] > df_v1['Limite_Filtro']) & \
                     (df_v1[col_exportador] != target_exporter)

    df_filtered_v1 = df_v1[~mask_delete_v1].copy()

    nombre_base = os.path.splitext(os.path.basename(file_path))[0]
    # Quitar el posible sufijo _CLASIFICADO si ya existe para evitar nombres excesivamente largos
    clean_nombre_base = nombre_base.replace("_CLASIFICADO", "")

    v1_path = os.path.join(folder_variations, f"{clean_nombre_base}_V1_Todos_Meses_Sin_Eliminar_PF.xlsx")
    save_filtered_excel(file_path, df_filtered_v1, original_cols, v1_path)

    # ----------------------------------------------------
    # 3. GENERAR VARIACIÓN 2 (Intervalo que contiene todas las ventas de Peru Frost, nunca elimina PF)
    # ----------------------------------------------------
    limites_v2 = {}
    for (mes, prod), group in df_full.groupby(['MES_ANIO', col_producto]):
        group_pf = group[group[col_exportador] == target_exporter]
        pf_prices = group_pf['Precio_Fila'].tolist()
        all_prices = group['Precio_Fila'].tolist()

        precio_ref = find_density_interval_price_with_pf(all_prices, pf_prices)
        if precio_ref is not None:
            limites_v2[(mes, prod)] = precio_ref + 0.5

    df_v2 = df_full.copy()
    df_v2['Limite_Filtro'] = df_v2.apply(lambda r: limites_v2.get((r['MES_ANIO'], r[col_producto])), axis=1)

    mask_delete_v2 = (df_v2['Limite_Filtro'].notna()) & \
                     (df_v2['Precio_Fila'] > df_v2['Limite_Filtro']) & \
                     (df_v2[col_exportador] != target_exporter)

    df_filtered_v2 = df_v2[~mask_delete_v2].copy()

    v2_path = os.path.join(folder_variations, f"{clean_nombre_base}_V2_Intervalo_Con_PF.xlsx")
    save_filtered_excel(file_path, df_filtered_v2, original_cols, v2_path)

    # ----------------------------------------------------
    # 4. GENERAR VARIACIÓN 3 (Criterio objetivo: IQR para mercado + todas las ventas de PF, nunca elimina PF)
    # ----------------------------------------------------
    limites_v3 = {}
    for (mes, prod), group in df_full.groupby(['MES_ANIO', col_producto]):
        group_pf = group[group[col_exportador] == target_exporter]
        group_others = group[group[col_exportador] != target_exporter]

        pf_fob = group_pf[col_fob_tot].sum()
        pf_kg = group_pf[col_kg_neto].sum()

        others_clean_fob = 0.0
        others_clean_kg = 0.0

        if not group_others.empty:
            prices_others = group_others['Precio_Fila'].values
            if len(prices_others) >= 4: # IQR tiene sentido con al menos 4 puntos
                q25 = pd.Series(prices_others).quantile(0.25)
                q75 = pd.Series(prices_others).quantile(0.75)
                iqr = q75 - q25
                lower_limit = q25 - 1.5 * iqr
                upper_limit = q75 + 1.5 * iqr

                group_others_clean = group_others[(group_others['Precio_Fila'] >= lower_limit) &
                                                  (group_others['Precio_Fila'] <= upper_limit)]
            else:
                group_others_clean = group_others

            others_clean_fob = group_others_clean[col_fob_tot].sum()
            others_clean_kg = group_others_clean[col_kg_neto].sum()

        total_fob = pf_fob + others_clean_fob
        total_kg = pf_kg + others_clean_kg

        if total_kg > 0:
            precio_real_promedio = total_fob / total_kg
            limites_v3[(mes, prod)] = precio_real_promedio + 0.5

    df_v3 = df_full.copy()
    df_v3['Limite_Filtro'] = df_v3.apply(lambda r: limites_v3.get((r['MES_ANIO'], r[col_producto])), axis=1)

    mask_delete_v3 = (df_v3['Limite_Filtro'].notna()) & \
                     (df_v3['Precio_Fila'] > df_v3['Limite_Filtro']) & \
                     (df_v3[col_exportador] != target_exporter)

    df_filtered_v3 = df_v3[~mask_delete_v3].copy()

    v3_path = os.path.join(folder_variations, f"{clean_nombre_base}_V3_Criterio_Objetivo.xlsx")
    save_filtered_excel(file_path, df_filtered_v3, original_cols, v3_path)

    # 5. Mostrar reporte resumido
    print(f"\n--- FILTRADO Y VARIACIONES COMPLETADAS CON ÉXITO!")
    print(f"   - Archivo Base: {os.path.basename(output_path)}")
    print(f"     (Filas: original={len(df_full)}, final={len(df_filtered_base)})")
    print(f"   - Variación 1 (Todos meses, sin eliminar PF): {os.path.basename(v1_path)}")
    print(f"     (Filas: final={len(df_filtered_v1)})")
    print(f"   - Variación 2 (Rango PF, sin eliminar PF): {os.path.basename(v2_path)}")
    print(f"     (Filas: final={len(df_filtered_v2)})")
    print(f"   - Variación 3 (IQR + PF, sin eliminar PF): {os.path.basename(v3_path)}")
    print(f"     (Filas: final={len(df_filtered_v3)})")


# ═══════════════════════════════════════════════════════════════════════════
# PARTE 3: PIPELINE UNIFICADO (clasifica y luego filtra, en un solo paso)
# ═══════════════════════════════════════════════════════════════════════════

def process_pipeline(file_path, classified_output_path, filtered_output_path):
    """
    Recibe un archivo Veritrade crudo, lo clasifica (columna AH = PRODUCTO)
    y encadena directamente el filtrado sobre el archivo recién clasificado.
    """
    try:
        process_veritrade(file_path, classified_output_path)
        process_filtering(classified_output_path, filtered_output_path)
    except Exception:
        print(f"[ERROR] procesando {os.path.basename(file_path)}:")
        traceback.print_exc()

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    input_dir = os.path.join(base_dir, "INPUT")
    output_dir = os.path.join(base_dir, "OUTPUT")
    filtered_dir = os.path.join(base_dir, "FILTERED_OUTPUT")

    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(filtered_dir, exist_ok=True)

    print(f"Buscando archivos *.xlsx en {input_dir}")
    archivos = glob.glob(os.path.join(input_dir, "*.xlsx"))

    if not archivos:
        print("TA VACIA la carpeta INPUT o tu archivo no es .xlsx CHECA ESO PORFA.")
        print("Pon los archivos Veritrade en la carpeta INPUT y vuelve a correr.")
        return

    for archivo in archivos:
        if os.path.basename(archivo).startswith("~$"):
            continue

        nombre_base = os.path.splitext(os.path.basename(archivo))[0]
        classified_output_path = os.path.join(output_dir, f"{nombre_base}_CLASIFICADO.xlsx")
        filtered_output_path = os.path.join(filtered_dir, f"{nombre_base}_CLASIFICADO_FILTRADO.xlsx")

        process_pipeline(archivo, classified_output_path, filtered_output_path)

    print("\n=====================================")
    print("TODO TERMINADO! Clasificación (col AH) + Filtrado completados.")

if __name__ == "__main__":
    main()
